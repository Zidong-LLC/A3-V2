"""Carga el correo de facturación de los clientes desde el Excel de Alegra (script one-off).

Uso:
    python tools/scripts/import_alegra_emails.py            # dry-run: solo reporta
    python tools/scripts/import_alegra_emails.py --apply    # escribe en clients.email

Cruza `Documentos de actualizacion/Alegra - Terceros v2 Actualizado.xlsx` contra la tabla
`clients` POR NIT (nunca por nombre: un match errado mandaría la factura de un cliente al
correo de otro). Reusa `db._nit_candidates()` para las variantes de dígito de verificación.

Solo completa lo que está vacío: si un cliente ya tiene correo cargado, no se pisa — el dato
de la base es más reciente que el de la planilla. Y solo escribe lo que valida como correo.

NO toca Alegra: esto prepara el dato en Supabase para que el flujo de facturación lo envíe.

Requiere `openpyxl` (no está en requirements.txt, el runtime no lo necesita):
    pip install openpyxl
"""
import argparse
import re
import sys
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))
load_dotenv(ROOT / ".env")

from app.services import db  # noqa: E402

EXCEL_PATH = ROOT / "Documentos de actualizacion" / "Alegra - Terceros v2 Actualizado.xlsx"
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[a-z]{2,}$", re.IGNORECASE)
COL_NAME, COL_NIT, COL_EMAIL = 0, 1, 8


def read_excel() -> list[dict]:
    workbook = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    rows = []
    for row in workbook.active.iter_rows(min_row=2, values_only=True):
        if not row[COL_NAME]:
            continue
        email = str(row[COL_EMAIL] or "").strip()
        if not EMAIL_RE.match(email):
            continue
        rows.append({"name": str(row[COL_NAME]).strip(), "tax_id": str(row[COL_NIT] or "").strip(), "email": email})
    workbook.close()
    return rows


def fetch_clients() -> list[dict]:
    out, offset = [], 0
    while True:
        page = db._client.table("clients").select("id, clinic_name, tax_id, email").range(offset, offset + 999).execute().data
        out += page
        if len(page) < 1000:
            break
        offset += 1000
    return out


def index_by_nit(clients: list[dict]) -> dict[str, list[dict]]:
    """Indexa cada cliente bajo todas las variantes de su NIT (con y sin dígito de verificación)."""
    index: dict[str, list[dict]] = {}
    for client in clients:
        for candidate in db._nit_candidates(client.get("tax_id") or ""):
            key = db._normalize_nit(candidate)
            if key:
                index.setdefault(key, []).append(client)
    return index


def main() -> int:
    parser = argparse.ArgumentParser(description="Carga clients.email desde el Excel de Alegra.")
    parser.add_argument("--apply", action="store_true", help="Escribe en la base (sin esto, solo reporta)")
    args = parser.parse_args()

    excel_rows = read_excel()
    clients = fetch_clients()
    index = index_by_nit(clients)

    to_write, already, unmatched = [], [], []
    multi_sede = 0
    for row in excel_rows:
        matches = []
        for candidate in db._nit_candidates(row["tax_id"]):
            for client in index.get(db._normalize_nit(candidate), []):
                if client not in matches:
                    matches.append(client)
        if not matches:
            unmatched.append(row)
            continue
        if len(matches) > 1:
            # Varias sedes con el mismo NIT son la MISMA veterinaria (y si es un veterinario
            # independiente, el NIT es el de su veterinaria). Alegra además crea un contacto
            # por NIT, así que el correo es por NIT: se carga en todas las sedes.
            multi_sede += 1
        for client in matches:
            if client.get("email"):
                already.append(row)
            else:
                to_write.append((row, client))

    print(f"Excel con correo válido : {len(excel_rows)}")
    print(f"  filas a escribir      : {len(to_write)}")
    print(f"  ya tenían correo      : {len(already)}")
    print(f"  NIT con varias sedes  : {multi_sede}  (se carga en todas: misma veterinaria)")
    print(f"  sin cliente por NIT   : {len(unmatched)}")

    if not args.apply:
        print("\nDRY-RUN. Nada se escribió. Volvé a correr con --apply para aplicar.")
        return 0

    for row, client in to_write:
        db._client.table("clients").update({"email": row["email"]}).eq("id", client["id"]).execute()
    print(f"\nEscritos: {len(to_write)} correos en clients.email")
    return 0


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    raise SystemExit(main())
