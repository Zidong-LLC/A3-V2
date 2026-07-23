# Carga las veterinarias del Excel "Clientes y Doctores A3.xlsx" (Hoja1) que no existen
# en `clients`, enriquecidas con NIT/dirección/teléfono de "Alegra - Terceros v2".
# Autorizado por el usuario (2026-07-22): cargarlas como clientes ACTIVOS para que el
# bot las identifique.
import re
import sys
import unicodedata

import openpyxl
from dotenv import load_dotenv

sys.path.insert(0, r"c:\Users\Artel\Downloads\A3 V2")
load_dotenv(r"c:\Users\Artel\Downloads\A3 V2\.env")
from app.services.db import _client  # noqa: E402

BASE = r"c:\Users\Artel\Downloads\A3 V2\Documentos de actualizacion"
APPLY = "--apply" in sys.argv


def key(s):
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]", "", s.lower())


# 1) Roster Hoja1
wb = openpyxl.load_workbook(rf"{BASE}\Clientes y Doctores A3.xlsx", read_only=True)
roster = {}
for nombre, _medico in wb["Hoja1"].iter_rows(min_row=2, values_only=True):
    if nombre and str(nombre).strip():
        roster.setdefault(key(nombre), str(nombre).strip())

# 2) Clientes actuales (todas las filas, activas o no)
clients = _client.table("clients").select("clinic_name").limit(10000).execute().data
db_keys = {key(c["clinic_name"]) for c in clients}

# 3) Faltantes = sin match exacto NI parcial (mismo criterio de la auditoría)
missing = {}
for k, nombre in roster.items():
    if k in db_keys:
        continue
    if any((len(k) >= 5 and k in kk) or (len(kk) >= 5 and kk in k) for kk in db_keys if kk):
        continue
    missing[k] = nombre

# 4) Alegra: enriquecer por nombre (columna "Clientes" trae "(Veterinaria X) Persona")
wb2 = openpyxl.load_workbook(rf"{BASE}\Alegra - Terceros v2 Actualizado.xlsx", read_only=True)
alegra = []
for row in wb2.worksheets[0].iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    full = str(row[0]).strip()
    paren = re.findall(r"\(([^)]+)\)", full)
    names = [p.strip() for p in paren] + [re.sub(r"\([^)]*\)", "", full).strip()]
    alegra.append({
        "names": [n for n in names if n],
        "tax_id": re.sub(r"[^0-9]", "", str(row[1] or "")),
        "address": str(row[4] or "").strip() or None,
        "phone": re.sub(r"[^0-9]", "", str(row[7] or row[5] or "")) or None,
    })


STOP = {"clinica", "veterinaria", "veterinario", "veterniario", "veterniaria", "centro",
        "consultorio", "vet", "sas", "sa", "de", "la", "el", "los", "las", "y", "dr", "dra",
        "doctor", "doctora", "sede"}


def toks(s):
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return {t for t in re.findall(r"[a-z0-9]+", s.lower()) if t not in STOP and len(t) >= 3}


def alegra_info(k, nombre):
    et = toks(nombre)
    candidates = []
    for a in alegra:
        for n in a["names"]:
            nk = key(n)
            at = toks(n)
            exact = nk and ((len(k) >= 5 and k in nk) or (len(nk) >= 5 and nk in k) or nk == k)
            # Subset solo con 2+ tokens en común: con uno solo ("friend") un NIT ajeno
            # se cuela en otra veterinaria — y un NIT equivocado es un error de facturación.
            subset = et and at and (et <= at or at <= et) and len(et & at) >= 2
            if exact or subset:
                candidates.append(a)
                break
    # Solo con match ÚNICO: ante ambigüedad no se inventa un NIT.
    return candidates[0] if len(candidates) == 1 else None


# phone es UNIQUE y NOT NULL: los clientes sin teléfono usan el placeholder '5700...'
# de la carga original. Continuamos la secuencia desde el máximo existente.
existing_phones = _client.table("clients").select("phone").like("phone", "5700%").limit(10000).execute().data
seq = max((int(p["phone"]) for p in existing_phones if str(p.get("phone") or "").isdigit()), default=570000000000)
used_phones = {str(c.get("phone")) for c in _client.table("clients").select("phone").limit(10000).execute().data}

rows, con_nit = [], 0
for k, nombre in sorted(missing.items(), key=lambda x: x[1].lower()):
    info = alegra_info(k, nombre) or {}
    if info.get("tax_id"):
        con_nit += 1
    phone = info.get("phone")
    if not phone or phone in used_phones:
        seq += 1
        phone = str(seq)
    used_phones.add(phone)
    rows.append({
        "clinic_name": nombre,
        "tax_id": info.get("tax_id") or None,
        # address es NOT NULL: vacío = "sin dato" (el bot pide la dirección de retiro
        # cuando la ficha no la tiene).
        "address": info.get("address") or "",
        "phone": phone,
        "billing_type": "cash",
        "is_active": True,
    })

print(f"Faltantes a cargar: {len(rows)} (con NIT desde Alegra: {con_nit})")
for r in rows:
    extra = f" | NIT {r['tax_id']}" if r["tax_id"] else ""
    extra += f" | {r['address']}" if r["address"] else ""
    print(f"  {r['clinic_name']}{extra}")

if not APPLY:
    print("\nDRY-RUN: no se escribió nada. Correr con --apply para insertar.")
    sys.exit(0)

# 5) Insertar en lotes de 50
inserted = 0
for i in range(0, len(rows), 50):
    batch = rows[i:i + 50]
    res = _client.table("clients").insert(batch).execute()
    inserted += len(res.data or [])
print(f"\nINSERTADOS: {inserted} clientes activos")
