"""Compara la cartera actualizada del cliente contra la tabla `clients` (solo lectura).

Uso:
    python tools/scripts/analyze_client_roster.py

NO escribe nada: clasifica y reporta para decidir con datos qué cargar. El match por nombre
exacto daba 266 "faltantes" y 403 "sobrantes", casi todos por variantes de escritura
("+Zoo Live", "AgrocolombiaSA"). Acá se usa el mismo matcher difuso que ya emplea el agente
para identificar clientes por nombre (`db._name_match_score`, umbral 0.85), más el NIT
cuando el Excel de Alegra lo aporta.

Regla dura: la ausencia de un cliente en la planilla NO prueba que dejó de serlo. Este script
nunca sugiere dar de baja a nadie.
"""
import re
import sys
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))
load_dotenv(ROOT / ".env")

from app.services import db  # noqa: E402

ROSTER_XLSX = ROOT / "Documentos de actualizacion" / "Clientes y Doctores A3.xlsx"
ALEGRA_XLSX = ROOT / "Documentos de actualizacion" / "Alegra - Terceros v2 Actualizado.xlsx"
MIN_SCORE = 1.6  # coverage 1.0 + ratio alto: exige todas las palabras y buen parecido global


def read_roster() -> list[str]:
    workbook = load_workbook(ROSTER_XLSX, read_only=True, data_only=True)
    names = {str(a).strip() for a, b in workbook["Hoja1"].iter_rows(min_row=2, values_only=True) if a}
    workbook.close()
    return sorted(names)


def read_alegra_nits() -> dict[str, dict]:
    """Nombre normalizado -> datos de Alegra (NIT, dirección, teléfono, correo)."""
    workbook = load_workbook(ALEGRA_XLSX, read_only=True, data_only=True)
    out = {}
    for row in workbook.active.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        out[db._normalize_lookup_key(row[0])] = {
            "tax_id": str(row[1] or "").strip(), "address": row[4],
            "phone": row[7] or row[5], "email": row[8],
        }
    workbook.close()
    return out


def fetch_clients() -> list[dict]:
    out, offset = [], 0
    while True:
        page = db._client.table("clients").select("id, clinic_name, tax_id, phone, address, email, is_active").range(offset, offset + 999).execute().data
        out += page
        if len(page) < 1000:
            break
        offset += 1000
    return out


def main() -> int:
    roster = read_roster()
    alegra = read_alegra_nits()
    clients = fetch_clients()

    by_nit = {}
    for client in clients:
        for candidate in db._nit_candidates(client.get("tax_id") or ""):
            key = db._normalize_nit(candidate)
            if key:
                by_nit.setdefault(key, []).append(client)
    by_key = {db._normalize_lookup_key(c["clinic_name"]): c for c in clients}

    exact, by_tax, fuzzy, missing = [], [], [], []
    for name in roster:
        key = db._normalize_lookup_key(name)
        if key in by_key:
            exact.append((name, by_key[key]))
            continue
        info = alegra.get(key) or {}
        hits = []
        for candidate in db._nit_candidates(info.get("tax_id") or ""):
            for client in by_nit.get(db._normalize_nit(candidate), []):
                if client not in hits:
                    hits.append(client)
        if hits:
            by_tax.append((name, hits[0]))
            continue
        tokens = [t for t in key.split("_") if t and t not in db._CLIENT_QUERY_STOPWORDS]
        compact = "".join(tokens)
        scored = sorted(
            ((db._name_match_score(tokens, compact, c["clinic_name"]), c) for c in clients),
            key=lambda pair: pair[0], reverse=True,
        )
        if scored and scored[0][0] >= MIN_SCORE:
            fuzzy.append((name, scored[0][1], round(scored[0][0], 2)))
        else:
            missing.append((name, info))

    print(f"Cartera del Excel        : {len(roster)} clínicas")
    print(f"  ya existen (nombre)    : {len(exact)}")
    print(f"  ya existen (por NIT)   : {len(by_tax)}")
    print(f"  ya existen (parecido)  : {len(fuzzy)}")
    print(f"  REALMENTE nuevas       : {len(missing)}")
    print()
    print("Ejemplos de match por parecido (verificar que no haya falsos positivos):")
    for name, client, score in fuzzy[:12]:
        print(f'   "{name[:32]:32}" -> "{client["clinic_name"][:34]:34}" ({score})')

    accionables = [(n, i) for n, i in missing if i.get("tax_id")]
    print()
    print(f"De las {len(missing)} nuevas: {len(accionables)} traen NIT (dan de alta), "
          f"{len(missing) - len(accionables)} solo el nombre (ficha de referencia).")
    for name, info in accionables[:10]:
        print(f'   {name[:38]:38} NIT={info["tax_id"]}')

    faltan_datos = [(n, c) for n, c in exact + by_tax if not c.get("email")]
    print()
    print(f"Clientes existentes SIN correo que el Excel podría completar: {len(faltan_datos)}")
    print("\nSOLO LECTURA: no se escribió nada.")
    return 0


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    raise SystemExit(main())
