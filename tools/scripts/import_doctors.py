"""Carga los médicos de la cartera actualizada en `clients_a3_professionals` (script one-off).

Uso:
    python tools/scripts/import_doctors.py            # dry-run: solo reporta
    python tools/scripts/import_doctors.py --apply    # escribe

La tabla ya existe con 1.828 filas de importaciones anteriores (ver su columna `source_sheet`).
Esto suma los médicos que la planilla nueva trae y no estaban.

`clinic_key` tiene FK contra `clients_a3_knowledge`, así que solo se cargan médicos de clínicas
con ficha: correr antes `import_client_roster.py --apply`. Los que queden fuera se reportan.
"""
import argparse
import sys
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))
load_dotenv(ROOT / ".env")

from app.services import db  # noqa: E402

ROSTER_XLSX = ROOT / "Documentos de actualizacion" / "Clientes y Doctores A3.xlsx"
SOURCE = "Clientes y Doctores A3 2026-07"


def read_pairs() -> list[tuple[str, str]]:
    workbook = load_workbook(ROSTER_XLSX, read_only=True, data_only=True)
    pairs = {(str(a).strip(), str(b).strip())
             for a, b in workbook["Hoja1"].iter_rows(min_row=2, values_only=True) if a and b}
    workbook.close()
    return sorted(pairs)


def fetch(table: str, columns: str) -> list[dict]:
    out, offset = [], 0
    while True:
        page = db._client.table(table).select(columns).range(offset, offset + 999).execute().data
        out += page
        if len(page) < 1000:
            break
        offset += 1000
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description="Carga médicos por clínica.")
    parser.add_argument("--apply", action="store_true", help="Escribe en la base (sin esto, solo reporta)")
    args = parser.parse_args()

    pairs = read_pairs()
    known_clinics = {row["clinic_key"] for row in fetch("clients_a3_knowledge", "clinic_key")}
    # Comparar por NOMBRE normalizado, no por `professional_key`: las cargas viejas lo
    # guardaron con espacios (y a veces el número de tarjeta pegado: "juan perez 4173 0"),
    # así que comparar claves crudas no detectaba al mismo médico y duplicaba filas.
    existing = {(row["clinic_key"], db._normalize_lookup_key(row.get("professional_name")))
                for row in fetch("clients_a3_professionals", "clinic_key, professional_name")}

    to_insert, no_clinic, already = [], [], 0
    for clinic, doctor in pairs:
        clinic_key = db._normalize_lookup_key(clinic)
        doctor_key = db._normalize_lookup_key(doctor)
        if not clinic_key or not doctor_key:
            continue
        if clinic_key not in known_clinics:
            no_clinic.append((clinic, doctor))
        elif (clinic_key, doctor_key) in existing:
            already += 1
        else:
            to_insert.append({"clinic_key": clinic_key, "professional_key": doctor_key,
                              "professional_name": doctor, "source_sheet": SOURCE})

    print(f"Pares clínica-médico en el Excel: {len(pairs)}")
    print(f"  ya cargados                   : {already}")
    print(f"  a insertar                    : {len(to_insert)}")
    print(f"  sin ficha de clínica (se omiten): {len(no_clinic)}")
    for clinic, doctor in no_clinic[:5]:
        print(f"     {clinic[:34]:34} / {doctor[:26]}")

    if not args.apply:
        print("\nDRY-RUN. Nada se escribió. Volvé a correr con --apply para aplicar.")
        return 0

    for start in range(0, len(to_insert), 500):
        db._client.table("clients_a3_professionals").upsert(
            to_insert[start:start + 500], on_conflict="clinic_key,professional_key,source_sheet"
        ).execute()
    total = db._client.table("clients_a3_professionals").select("id", count="exact").limit(1).execute().count
    print(f"\nInsertados: {len(to_insert)} | total en la tabla: {total}")
    return 0


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    raise SystemExit(main())
