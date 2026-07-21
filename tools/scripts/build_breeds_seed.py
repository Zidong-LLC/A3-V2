"""Genera el seed SQL del catálogo de RAZAS desde el Excel del cliente (script one-off).

Uso:
    python tools/scripts/build_breeds_seed.py

Lee `Documentos de actualizacion/Lista de Especies con Raza.xlsx` (columna A = especie,
columna B = raza) y escribe `db/seeds/003_catalog_breeds.sql`. El SQL se commitea: este
script solo se vuelve a correr si el cliente manda una lista nueva.

El mapeo de nombres de especie del Excel a los canónicos de A3 vive SOLO acá — `app/breeds.py`
no traduce nada. Si el Excel trae una especie que `app/species.py` no conoce, el script aborta:
no puede introducir vocabulario nuevo en silencio.

Requiere `openpyxl` (no está en requirements.txt porque el runtime no lo necesita):
    pip install openpyxl
"""
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from app.species import RECOVERABLE_SPECIES  # noqa: E402
from app.text import catalog_item_key  # noqa: E402

EXCEL_PATH = ROOT / "Documentos de actualizacion" / "Lista de Especies con Raza.xlsx"
OUTPUT_PATH = ROOT / "db" / "seeds" / "003_catalog_breeds.sql"

# Nombres del Excel → canónicos de app/species.py. Las especies que ya coinciden
# (Canino, Felino, Bovino, Porcino, Equino, Ovino, Caprino, Hurón) no se listan.
EXCEL_SPECIES_TO_A3 = {
    "Aviar": "Ave",
    "Lagomorfo": "Conejo",
    "Cobayo": "Roedor",
    "Hámster": "Roedor",
    "Reptiles": "Reptil",
}
# La hoja "Exóticos" no trae razas sino especies sueltas: cada una es su propia especie.
EXOTIC_BREED_TO_SPECIES = {
    "Erizo Africano": "Erizo",
    "Chinchilla": "Chinchilla",
    "Sugar Glider": "Sugar Glider",
    "Degú": "Degú",
    "Axolote": "Axolote",
}


def read_rows() -> list[tuple[str, str]]:
    """Devuelve (especie_canónica, raza) por cada fila del Excel, sin deduplicar."""
    workbook = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    rows: list[tuple[str, str]] = []
    for sheet in workbook.worksheets:
        for species_cell, breed_cell in sheet.iter_rows(min_col=1, max_col=2, values_only=True):
            if not species_cell or not breed_cell:
                continue
            breed = str(breed_cell).strip()
            if breed.lower() == "raza":  # fila de encabezado (solo la hoja Canino la trae)
                continue
            excel_species = str(species_cell).strip()
            if excel_species == "Exóticos":
                species = EXOTIC_BREED_TO_SPECIES[breed]
            else:
                species = EXCEL_SPECIES_TO_A3.get(excel_species, excel_species)
            rows.append((species, breed))
    workbook.close()
    return rows


def build_catalog(rows: list[tuple[str, str]]) -> list[tuple[str, str, str]]:
    """Deduplica por (especie, clave de raza) y ordena para que el diff sea estable."""
    seen: dict[tuple[str, str], tuple[str, str, str]] = {}
    for species, breed in rows:
        key = catalog_item_key(breed)
        if (species, key) not in seen:
            seen[(species, key)] = (key, breed, species)
    return sorted(seen.values(), key=lambda item: (item[2], item[1]))


def report(rows: list[tuple[str, str]], catalog: list[tuple[str, str, str]]) -> None:
    by_key: dict[str, set[str]] = defaultdict(set)
    for key, _breed, species in catalog:
        by_key[key].add(species)
    ambiguous = sorted(k for k, v in by_key.items() if len(v) > 1)
    collisions = sorted(k for k in by_key if k in RECOVERABLE_SPECIES)
    print(f"Filas leídas del Excel : {len(rows)}")
    print(f"Filas del catálogo     : {len(catalog)}")
    print(f"Razas únicas           : {len(by_key)}")
    print(f"Ambiguas entre especies: {len(ambiguous)} -> {', '.join(ambiguous)}")
    print(f"Colisionan con especie : {len(collisions)} -> {', '.join(collisions)}")


def assert_species_are_canonical(catalog: list[tuple[str, str, str]]) -> None:
    canonical = set(RECOVERABLE_SPECIES.values())
    unknown = sorted({species for _key, _breed, species in catalog} - canonical)
    if unknown:
        raise SystemExit(
            f"ABORTA: especies que app/species.py no conoce: {', '.join(unknown)}.\n"
            "Agregalas a ANIMAL_DOMAIN o corregí EXCEL_SPECIES_TO_A3."
        )


def render_sql(catalog: list[tuple[str, str, str]]) -> str:
    lines = [
        "-- Ejecutar en el SQL Editor de Supabase después de 016_catalog_breeds.sql",
        "-- Catálogo de RAZAS por especie — lista entregada por el cliente",
        '-- Fuente: Excel "Lista de Especies con Raza" (7 hojas, columna A = especie, B = raza)',
        "-- Generado por tools/scripts/build_breeds_seed.py — no editar a mano",
        "",
        "INSERT INTO catalog_breeds (breed_key, name, species) VALUES",
    ]
    width = max(len(f"'{key}',") for key, _breed, _species in catalog)
    current_species = None
    values: list[str] = []
    for key, breed, species in catalog:
        if species != current_species:
            current_species = species
            header = f"-- ── {species.upper()} "
            values.append(f"\n{header}{'─' * max(0, 78 - len(header))}")
        key_field = f"'{key}',".ljust(width)
        values.append(f"({key_field} '{breed.replace(chr(39), chr(39) * 2)}', '{species}'),")
    body = "\n".join(values).rstrip(",")
    lines.append(body.lstrip("\n"))
    lines += [
        "",
        "ON CONFLICT (breed_key, species) DO UPDATE SET",
        "    name      = EXCLUDED.name,",
        "    is_active = true;",
        "",
    ]
    return "\n".join(lines)


def main() -> None:
    rows = read_rows()
    catalog = build_catalog(rows)
    assert_species_are_canonical(catalog)
    report(rows, catalog)
    OUTPUT_PATH.write_text(render_sql(catalog), encoding="utf-8")
    print(f"\nEscrito: {OUTPUT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
