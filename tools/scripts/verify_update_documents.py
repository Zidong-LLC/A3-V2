"""Verificador de "Documentos de actualizacion" contra Supabase.

Correr DESPUÉS de cada actualización de los Excel de clientes/razas/Alegra:

    python tools/scripts/verify_update_documents.py

Chequea que TODO lo que está en los documentos exista en la base:
  1. Veterinarias (Hoja1 de "Clientes y Doctores A3.xlsx") -> clients
  2. Razas ("Lista de Especies con Raza.xlsx") -> catalog_breeds (por nombre)
  3. Terceros de Alegra (NIT) -> clients.tax_id
  4. Médicos (Hoja1) -> clients_a3_professionals
Sale con código != 0 si falta algo (sirve de gate en CI o antes de un deploy).

Nació del hueco 2026-07-22: la carga anterior dejó 188 veterinarias del roster en
knowledge/professionals pero NO en `clients`, y el bot no las identificaba (caso
"AgrocolombiaSA"). Regla: ninguna actualización de documentos se da por cerrada sin
este verificador en cero.
"""
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))
load_dotenv(ROOT / ".env")

from app.services import db  # noqa: E402
from app.services.db import _client  # noqa: E402

DOCS = ROOT / "Documentos de actualizacion"

# Etiquetas del Excel de razas -> especie canónica del dominio del bot (app/species.py)
SPECIES_LABEL_MAP = {
    "aviar": "Ave", "lagomorfo": "Conejo", "hamster": "Roedor", "cobayo": "Roedor",
    "reptiles": "Reptil", "exoticos": None,  # exóticos: cada raza es su propia especie
}


def key(s):
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]", "", s.lower())


def fetch_all(table, select):
    out, page = [], 0
    while True:  # PostgREST trunca en 1.000 filas: SIEMPRE paginar
        batch = (_client.table(table).select(select)
                 .range(page * 1000, page * 1000 + 999).execute().data)
        out += batch
        if len(batch) < 1000:
            return out
        page += 1


def check_clinics() -> list[str]:
    wb = openpyxl.load_workbook(DOCS / "Clientes y Doctores A3.xlsx", read_only=True)
    roster = {key(n): str(n).strip() for n, _ in wb["Hoja1"].iter_rows(min_row=2, values_only=True)
              if n and str(n).strip()}
    db_keys = {key(c["clinic_name"]) for c in fetch_all("clients", "clinic_name")}
    missing = []
    for k, nombre in roster.items():
        if k in db_keys:
            continue
        if any((len(k) >= 5 and k in kk) or (len(kk) >= 5 and kk in k) for kk in db_keys if kk):
            continue  # variante de nombre ya registrada (sede/prefijo)
        # Último recurso: el matching REAL del bot (cubre typos del Excel, p. ej.
        # "Venencia" vs "Venecia"). Si el bot lo encuentra, el cliente está operativo.
        if db.find_client_exact(nombre) or db.find_client_matches(nombre, limit=1):
            continue
        missing.append(nombre)
    return missing


def check_breeds() -> list[str]:
    wb = openpyxl.load_workbook(DOCS / "Lista de Especies con Raza.xlsx", read_only=True)
    excel = set()
    for ws in wb.worksheets:
        for _sp, raza in ws.iter_rows(values_only=True):
            if raza and str(raza).strip().lower() != "raza":
                excel.add(str(raza).strip())
    db_names = {key(b["name"]) for b in fetch_all("catalog_breeds", "name")}
    return [r for r in sorted(excel) if key(r) not in db_names]


def check_alegra() -> list[str]:
    wb = openpyxl.load_workbook(DOCS / "Alegra - Terceros v2 Actualizado.xlsx", read_only=True)
    ws = wb.worksheets[0]
    clients = fetch_all("clients", "clinic_name,tax_id")
    db_nits = set()
    for c in clients:
        digits = re.sub(r"[^0-9]", "", str(c.get("tax_id") or ""))
        if digits:
            db_nits.add(digits.lstrip("0"))
            db_nits.add(digits.lstrip("0")[:-1])  # sin dígito de verificación/sufijo de sede
    missing = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:
            continue
        nit = re.sub(r"[^0-9]", "", str(row[1])).lstrip("0")
        if nit and nit not in db_nits and nit[:-1] not in db_nits:
            missing.append(f"{str(row[0]).strip()} (NIT {nit})")
    return missing


def check_doctors() -> list[str]:
    wb = openpyxl.load_workbook(DOCS / "Clientes y Doctores A3.xlsx", read_only=True)
    excel_docs = {key(m): str(m).strip() for _, m in wb["Hoja1"].iter_rows(min_row=2, values_only=True)
                  if m and str(m).strip()}
    db_names = {key(p["professional_name"]) for p in
                fetch_all("clients_a3_professionals", "professional_name")}
    missing = []
    for k, nombre in excel_docs.items():
        if k in db_names:
            continue
        if any((len(k) >= 8 and k in kk) or (len(kk) >= 8 and kk in k) for kk in db_names if kk):
            continue  # variante con/sin "Dr." o segundo apellido
        missing.append(nombre)
    return missing


def main() -> int:
    fallas = 0
    for titulo, faltantes in (
        ("VETERINARIAS (Hoja1 -> clients)", check_clinics()),
        ("RAZAS (Lista de Especies -> catalog_breeds)", check_breeds()),
        ("TERCEROS ALEGRA (NIT -> clients.tax_id)", check_alegra()),
        ("MÉDICOS (Hoja1 -> clients_a3_professionals)", check_doctors()),
    ):
        estado = "OK — todo cargado" if not faltantes else f"FALTAN {len(faltantes)}"
        print(f"[{estado}] {titulo}")
        for f in faltantes:
            print(f"    - {f}")
        fallas += len(faltantes)
    print(f"\n{'TODO EN ORDEN' if not fallas else f'TOTAL FALTANTES: {fallas}'}")
    return 0 if not fallas else 1


if __name__ == "__main__":
    raise SystemExit(main())
